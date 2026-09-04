"""Generate "FSAE-Sim Parameters.xlsx" — the team-facing parameter sheet.

Three tabs:
  1. Car parameters   — GENERATED from the params.yaml files. Every column
                        (label, symbol, what it is, value, units, provenance,
                        what we measure, how, why it matters) is a field of
                        the YAML entry, so the sheet cannot drift from the sim
                        and adding a parameter never means editing this file.
  2. Sim signals      — the quantities the sim COMPUTES each step (slip
                        angle, slip ratio, loads, yaw rate, ...): what they
                        are, how the sim gets them, and which sensor sees
                        them on the real car.
  3. Test values      — the maneuver knobs and the real envelope numbers.

Tabs 2 and 3 are editorial, not parameter data, so they stay hand-written
below. Tab 1 used to be hand-written too — a 250-line table that duplicated
every description already in car_data.py's comments and had to be kept in
sync by hand. It is now four loops over cfg.

Adding a parameter to the sheet: add it to the relevant params.yaml. Adding a
whole SECTION: give the new params.yaml an `order:` and a `title:`.

Color rule (team convention): values confirmed for the CURRENT CAR or
DRIVER are GREEN; everything else (report car, guesses, rules, tuned gains)
is black. When a value is later MEASURED it stays green.

Rebuild after editing any params.yaml:
    .venv/bin/python param_sheet.py

Import to Google Sheets: drag the .xlsx into drive.google.com — formatting
and colors carry over.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from model.config import cfg, GREEN_TAGS

OUT = os.path.join("docs", "datasheets", "FSAE-Sim Parameters.xlsx")

GREEN = "FF1A7A1A"          # current-car / driver / measured values
BLACK = "FF1A1A1A"
HDR_FILL = PatternFill("solid", fgColor="FF223038")
SEC_FILL = PatternFill("solid", fgColor="FFE8EAE4")
WARN = "FF9A6A00"           # placeholder annotation text

THIN = Side(style="thin", color="FFD9DBD4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fmt(v):
    """Human-friendly number."""
    if isinstance(v, float):
        if v == 0:
            return "0"
        if abs(v) >= 1000:
            return f"{v:,.0f}"
        return f"{v:.4g}"
    return str(v)


def field(entry, name, default="—"):
    """One documentation field of a parameter, as a display string."""
    v = entry.get(name)
    if v is None or (isinstance(v, str) and not v.strip()):
        return default
    return " ".join(str(v).split())      # collapse YAML block-scalar wrapping


def units_of(entry):
    """What to show in the Units column: the human form when the entry gives
    one (`kg (weighed as 400.1 lb)`), otherwise the machine unit."""
    return field(entry, "display_unit", default=field(entry, "unit"))


# ──────────────────────────────────────────────────────────────────────
# Tab 2 — computed signals
# ──────────────────────────────────────────────────────────────────────
S = [
 ("Slip angle", "α", "alpha (vehicle.py)",
  "Angle between where a wheel POINTS and where its contact patch actually TRAVELS. The input that generates lateral force.",
  "Computed every step per wheel: α = −atan2(v_lateral, v_forward) at the contact patch",
  "Not directly measurable cheaply — estimated from IMU + steering angle; optical sensors (Kistler) exist but cost more than the car",
  "Every tire always runs some slip angle — that IS how it grips. Past ~11.6° (placeholder peak) more angle = less force."),
 ("Slip ratio", "κ", "kappa (vehicle.py)",
  "How much faster the wheel SURFACE moves than the ground: κ = (ω·r_w − v)/v. Positive = wheelspin side.",
  "Computed per driven wheel from its spin state and contact-patch speed",
  "Motor resolver speed (AMK reports it over CAN) ÷ gear ratio vs estimated ground speed",
  "The s-diff's whole world. Past the peak (~0.10) the wheel runs away — that's wheelspin, and it emerges from the ODE, no if-statement."),
 ("Wheel spin speeds", "ω_RL, ω_RR", "states 7–8",
  "Rear wheel rotational speeds — real states with their own differential equation I·ω̇ = T − r·Fx.",
  "Integrated by RK4 like everything else",
  "AMK motor resolvers over CAN (÷ gear ratio) — free, accurate, fast",
  "The signal the real s-diff will actually control. Sensor already exists in the motors."),
 ("Yaw rate", "r", "state 6",
  "How fast the car rotates about vertical. Left turn positive.",
  "Integrated from Iz·ṙ = ΣMz",
  "IMU gyro (any automotive-grade IMU; the VCU needs one anyway)",
  "Sets the s-diff wheel-speed target (r·track/r_wheel). Its sensor is a non-negotiable hardware item."),
 ("Body velocities", "vx, vy", "states 4–5",
  "Forward and sideways speed of the CG in the car's own frame.",
  "Integrated from the force balance",
  "vx: wheel speeds + IMU fusion (or GPS); vy: estimated (observer) — no cheap direct sensor",
  "vx feeds the slip math; vy is why sideslip needs estimation on the real car."),
 ("Body sideslip", "β", "beta (logged)",
  "Angle between where the car POINTS and where it TRAVELS: β = atan(vy/vx). Big β = the rear stepping out.",
  "Computed from vy/vx",
  "Estimated (IMU + model observer); optical ground-speed sensors measure it directly but are exotic",
  "The honest 'is the car sliding' number."),
 ("Vertical loads", "Fz (×4)", "wheel_loads() (vehicle.py)",
  "Weight on each tire right now: static + downforce + longitudinal & lateral load transfer.",
  "Algebraic from ax, ay, speed each step (quasi-static, 3-iteration loop)",
  "Suspension load cells or pushrod strain gauges (nice-to-have); corner scales give the static part",
  "Grip is proportional to load (minus load sensitivity) — every force the tire model produces starts from Fz."),
 ("Tire forces", "Fx, Fy (×4)", "tire_forces() (vehicle.py)",
  "What each contact patch pushes on the ground: drive/brake (Fx) and cornering (Fy).",
  "Magic Formula from (α, κ, Fz), capped by the friction circle √(Fx²+Fy²) ≤ µFz",
  "Not directly measurable — inferred from IMU accelerations and wheel torques",
  "The only place the car touches the world. Everything else is bookkeeping around these eight numbers."),
 ("Friction utilization", "—", "verify.py audit",
  "How much of a tire's total grip budget is in use: √(Fx²+Fy²)/(µFz). 1.00 = on the limit.",
  "Computed in the audit/replay from the forces",
  "Inferred (same as forces)",
  "The corner-exit test runs the rear axle at 1.00 — meaning results live on the guessed part of the tire curve."),
 ("Accelerations", "ax, ay", "logged",
  "Specific forces on the body (what an accelerometer strapped to the car would read).",
  "ΣF/m each step",
  "IMU accelerometer — cheap, already needed",
  "ay·vx-vs-yaw-rate cross-checks (ay = r·vx steady state) are how sim vs real-car validation will start."),
 ("Yaw moment", "Mz", "derivatives() (vehicle.py)",
  "Net twisting moment about the CG from all eight tire force components.",
  "Mz = Σ(x·Fy − y·Fx) over the four wheels",
  "Not measurable directly — its EFFECT (yaw acceleration) is, via the gyro",
  "(track/2)·ΔFx of it comes from the torque split."),
 ("Wheel-speed-difference target", "Δω_target", "controllers.py",
  "The rear-wheel speed difference corner geometry requires: r·track/r_wheel.",
  "Computed from measured yaw rate each controller update",
  "Derived from the gyro on the real car",
  "The s-diff's setpoint: allow exactly this difference, fight spin beyond it."),
 ("Torque split", "ΔT", "controllers.py",
  "The left/right torque difference: ΔT = ΔT_sdiff, then limits applied.",
  "Controller output every update",
  "Commanded over CAN to the AMK inverters; they report actual torque back",
  "The single actuator this whole project exists to command."),
 ("Wheel torques", "T_RL, T_RR", "controllers.py output",
  "Final per-wheel torque commands after every limit (peak, regen, power caps).",
  "Base ± ΔT/2, then the limits chain",
  "AMK inverters accept torque setpoints and report actuals over CAN",
  "What actually gets sent. The limits chain can silently shrink ΔT at high speed — log requested vs delivered on the car."),
 ("Total drive power", "P", "logged",
  "Mechanical power at the wheels: T_RL·ω_RL + T_RR·ω_RR.",
  "Computed from torques × speeds",
  "Rules energy meter measures the electrical version at the accumulator",
  "The 80 kW rules cap lives here (sim: mechanical; real: electrical — real car hits it ~10–15% sooner)."),
]

# ──────────────────────────────────────────────────────────────────────
# Tab 3 — the tests: what each measures, how, and the numbers we want
# ──────────────────────────────────────────────────────────────────────
T = [
 ("Step steer", "5° step at 15 m/s, torque holds speed",
  "How the car answers a sudden steering input: rise time, overshoot, steady-state accuracy.",
  "Straight running, then a fast (~0.1 s) ramp to a fixed steer angle, held to the end.",
  "NO overshoot/oscillation in the yaw trace (oscillation = gains too hot). Sideslip β < ~1°.",
  "All configs stable."),
 ("Corner exit (power on)", "8° corner at 10 m/s, throttle ramp to 45% of peak, unwind",
  "Traction management — the s-diff's reason to exist: put power down mid-corner without spinning the unloaded inner wheel or the car.",
  "Turn in, steady corner, then throttle ramps in exactly when the inner rear is lightest, wheel unwinding on exit.",
  "MAX SLIP RATIO κ on the inner rear — the headline. Must stay under the tire peak κ≈0.10. At the 23° envelope test: open diff κ=67 (violent spin) vs s-diff κ=0.07 — that single pair of numbers IS the s-diff justification. Also: no spin flag, yaw tracked while power is down.",
  "s-diff configs contain κ; open shows the failure. NOTE: with real R20 grip the 45% default barely stresses the axle — run --corner-throttle 65 (or the 23° envelope) to separate configs."),
 ("Slalom (sine steer)", "±4° at 0.5 Hz, 15 m/s",
  "Transient handling through repeated direction changes: phase lag, left/right symmetry, and whether errors GROW cycle-to-cycle (the spin precursor).",
  "Sine-wave steering, amplitude faded in over 1 s, constant drive torque.",
  "β bounded and symmetric; amplitude NOT growing cycle to cycle (growth = impending spin); Δω flips sign cleanly each reversal.",
  "Nothing diverges at the default point."),
 ("Pedal check (APPS/BPS)", "APPS 60% → overlap with brake → release → regen",
  "The pedal→torque chain and the FSAE EV.4.7 plausibility cut. Rules compliance, not handling — pass/fail.",
  "Straight line: throttle, then brake pressure WHILE throttle held (illegal overlap), release both, then brake alone.",
  "Wheel torque EXACTLY 0 through the whole overlap (run 015: 0.000 ✓); cut latched until APPS < 5%; regen negative and inside T_REGEN_MAX; vx estimate tracks through it.",
  "Hard pass/fail — any nonzero torque in the overlap is a rules violation."),
 ("Envelope: max steer", "23° road-wheel (CURRENT CAR)",
  "The physical steering lock — robustness test point, not an operating point.",
  "Run any maneuver at 23° via ./run.sh prompts or flags.",
  "Full-lock hairpin: the s-diff κ-containment headline. Full-lock slalom at 40 mph: which configs survive.",
  "β stays bounded."),
 ("Envelope: top speed", "40 mph = 17.9 m/s (CURRENT CAR)",
  "Endurance top expected speed.",
  "Speed test point for step/slalom.",
  "Above ~17 m/s the 20 kW motor power cap trims torque (256 of 273 N·m at 40 mph) — check commands respect it.",
  "—"),
 ("Changing test points", "./run.sh asks · flags · maneuvers.py defaults",
  "Every run RECORDS its test values; changing one auto-labels the run and flags 'metrics not comparable' vs the previous run.",
  "—", "—", "—"),
]


# ──────────────────────────────────────────────────────────────────────
# build the workbook
# ──────────────────────────────────────────────────────────────────────
def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFFFF", size=11)
        cell.fill = HDR_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER


def add_sheet(wb, title, headers, widths):
    ws = wb.create_sheet(title)
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    return ws


wb = Workbook()
wb.remove(wb.active)

# ---- tab 1 — generated from the params.yaml files
H1 = ["Parameter", "Symbol", "Name in the code", "What it is / what it simulates",
      "Value now", "Units", "Status (where the number comes from)",
      "What we measure", "How we measure it", "Why it matters"]
W1 = [26, 9, 30, 46, 13, 16, 34, 30, 38, 48]
ws = add_sheet(wb, "Car parameters", H1, W1)

r = 2
for namespace, sec in cfg.section_list():
    ws.cell(row=r, column=1, value=sec["title"])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(H1))
    c = ws.cell(row=r, column=1)
    c.font = Font(bold=True, size=10.5, color=BLACK)
    c.fill = SEC_FILL
    r += 1

    for path, e in cfg.section_params(namespace).items():
        status = field(e, "status", default="—")
        tag = cfg.tag_of(path)
        green = tag in GREEN_TAGS
        # A derived value shows its formula where an entered value would show
        # its source, so the sheet says how it is computed, not just what it is.
        how = (f"= {field(e, 'derived')}" if "derived" in e
               else field(e, "how"))
        vals = [field(e, "label", default=e["name"]),
                field(e, "symbol", default="—"),
                path,
                field(e, "what"),
                fmt(e["si"]),
                units_of(e),
                status,
                field(e, "need"),
                how,
                field(e, "why")]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if ci == 5:                     # the value column
                cell.font = Font(color=GREEN if green else BLACK, bold=green)
            elif ci == 7:                   # the status column
                color = (GREEN if green else
                         WARN if ("PLACEHOLDER" in status.upper()
                                  or "SUSPECT" in status.upper()) else BLACK)
                cell.font = Font(color=color, size=10)
            else:
                cell.font = Font(color=BLACK, size=10)
        r += 1

# legend
r += 1
ws.cell(row=r, column=1, value="Legend:").font = Font(bold=True)
ws.cell(row=r, column=2, value="GREEN value = confirmed for the CURRENT car/driver (or MEASURED). "
        "Black = report car, guess, rules, or sim-tuned. Orange status = placeholder/needs attention. "
        "Every row is generated from a params.yaml file — edit the number there, not here.")
ws.cell(row=r, column=2).font = Font(color=BLACK, size=10)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=len(H1))

# ---- tab 2
H2 = ["Signal", "Symbol", "Where in the code", "What it is",
      "How the sim gets it", "How we'd measure it on the real car", "Why it matters"]
W2 = [24, 12, 24, 48, 42, 44, 52]
ws2 = add_sheet(wb, "Sim signals", H2, W2)
r = 2
for row in S:
    for ci, v in enumerate(row, start=1):
        cell = ws2.cell(row=r, column=ci, value=v)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER
        cell.font = Font(color=BLACK, size=10)
    r += 1

# ---- tab 3
H3 = ["Test", "Current test point", "What it measures", "How it works",
      "The numbers we want (and what good looks like)", "Expected outcome"]
W3 = [22, 30, 44, 40, 60, 40]
ws3 = add_sheet(wb, "Tests", H3, W3)
r = 2
for row in T:
    green = "CURRENT CAR" in row[1]
    for ci, v in enumerate(row, start=1):
        cell = ws3.cell(row=r, column=ci, value=v)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER
        if ci == 2:
            cell.font = Font(color=GREEN if green else BLACK, bold=green, size=10)
        else:
            cell.font = Font(color=BLACK, size=10)
    r += 1

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
n_params = sum(len(cfg.section_params(ns)) for ns, _ in cfg.section_list())
print(f"wrote {OUT}  ({n_params} parameters in "
      f"{len(cfg.section_list())} sections, generated from the params.yaml files)")
print("Import to Google Sheets: drag the file into drive.google.com — "
      "colors and layout carry over.")
